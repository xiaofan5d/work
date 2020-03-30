# -*- coding: utf-8 -*-
'''
---------------------------------------------------------------------
具体步骤:
1.导入库存,订单以及厂价表
2.库存需要维护厂价表(订单中有的SKU厂价表里没有)
(需要考虑码对应问题)
3.订单的sku库存无需要报错(具体用商品编码对应商品代码匹配)
4.开始QA-散货发货
5.开始QA-整箱发货
6.开始其他-散货发货
7.开始其他-整箱发货
--------------------------------------------------------------------
修改分货逻辑
1 订单拆分
（1）采购数量为0
（2）QA散货分货
（3）QA整箱分货
（4）其他散货分货
（5）其他整箱分货
2 订单以效期先入先出原则为准
订单拆分后包含效期
3 具体分货流程为排序效期 计算累加和 再计算分货数量(需要考虑货不足的情况发生)
4 重要:订单拆分过程中如果没数据需要报错或者跳过
（1）库存长度为0需要跳出进入下一个
（2）每次订单长度为0需要跳出进入下一个

P.S
开发gui界面
箱规原则上先筛选目前有的库存的箱规 再以min箱规为准,无数据的箱规数据改为1,
并报错(某某sku无箱规,默认填为1)

'''
import math
from datetime import timedelta
from datetime import datetime
from pandas import *
import numpy
import openpyxl
from openpyxl.styles import Font, Alignment ,Border,Side,PatternFill
import excel_cl
#%%
#库存处理
def stock_cl(stock,sku,Order):
    x = stock[(stock.货物状态 == 'NR') | (
            stock.货物状态 == 'OD') | (stock.货物状态 == 'QA')]
    x = pandas.merge(left=x,
                     right=sku.loc[:,
                           ['公司货号',
                            '商品编码',
                            '条码',
                            '保质期',
                            '体积',
                            '重量',
                            '箱规']].drop_duplicates(),
                     left_on='商品代码',
                     right_on='公司货号',
                     how='left').fillna('dummy')
    x = x.pivot_table(
        index=[
            '商品编码',
            '条码',
            '商品代码',
            '商品名称',
            '库位',
            '生产批号',
            '到期日期',
            '箱规',
            '保质期',
            '体积',
            '重量',
            '货物状态'],
        values=['可用数量'],
        aggfunc='sum').reset_index().replace('dummy',numpy.nan)
    z=pandas.Series(Order.商品编码.unique())
    x=x[x.商品编码.isin(z) == True]
    x['散货数量'] = x.可用数量 - (x.可用数量 / x.箱规).astype(int) * x.箱规
    x['可用数量'] = x.可用数量 - x.散货数量
    库存拆分1 = x.drop('散货数量', axis=1)
    库存拆分2 = x.drop('可用数量', axis=1).rename(columns={'散货数量': '可用数量'})
    库存拆分1 = 库存拆分1[库存拆分1.可用数量 != 0]
    库存拆分1['箱规属性'] = '整箱'
    库存拆分2 = 库存拆分2[库存拆分2.可用数量 != 0]
    库存拆分2['箱规属性'] = '散装'
    库存拆分 = [库存拆分1, 库存拆分2]
    库存 = pandas.concat(库存拆分)
    return 库存
def Order_cl(stock,Order,sku):
    z1 = pandas.Series(Order.商品编码.unique())
    z2 = stock.loc[stock.商品编码.isin(z1) == True, ['商品编码', '箱规']].drop_duplicates()
    x = pandas.merge(left=Order,
                     right=z2,
                     on='商品编码',
                     how='left')
    if len(x.loc[x.箱规.isnull()==True])>1:
        print('\033[1;31m 警告:有存在SKU库存无货,具体明细为: \033[0m')
        print(x.loc[x.箱规.isnull() == True,['商品编码','商品名称']])
        x = pandas.merge(left=x,
                     right=sku.loc[:, ['商品编码', '箱规']].drop_duplicates(),
                     on='商品编码',
                     how='left')
        x.loc[x.箱规_x.isnull()==True,'箱规_x']=x.loc[x.箱规_x.isnull()==True,'箱规_y']
        x=x.rename(columns={'箱规_x':'箱规'})
    z=x.loc[:, ['商品编码', '箱规']].drop_duplicates().商品编码.value_counts()
    if len(z[z.values>1])>1:
        print('\033[1;31m 警告:此次订单有SKU有多个箱规 \033[0m')
        return z[z.values > 1].index
    else:
        print('\033[1;32m 正常:sku箱规正常(一对一) \033[0m')
    return x
##判断sku是否在厂价表中的函数
def isin_sku(Order,sku):
    z=pandas.Series(Order.商品编码.unique())
    z1=pandas.Series(sku.商品编码.unique())
    if len(z[z.isin(z1)==False])>0:
        print('\033[1;31m 警告:订单采购SKU需要维护,具体明细为: \033[0m')
        return z[z.isin(z1) == False]
    else:
        print('\033[1;32m 正常:订单sku正常 \033[0m')
def Order_new(订单处理_最终,订单_bak):
    订单处理_1 = 订单处理_最终.loc[订单处理_最终.仓库实发数量 != 0].pivot_table(
        index=['订单号', '商品编码'], values=['仓库实发数量'], aggfunc='sum').reset_index()
    订单_new = pandas.merge(
        left=订单_bak, right=订单处理_1, left_on=[
            '商品编码', '订单号'], right_on=[
            '商品编码', '订单号'], how='left')
    if len(订单处理_1) == 0:
        订单_new['采购数量1'] = 订单_new.采购数量1 - 0
    else:
        订单_new['仓库实发数量'] = 订单_new.仓库实发数量.fillna(0)
        订单_new['采购数量1'] = 订单_new.采购数量1 - 订单_new.仓库实发数量
    return 订单_new.loc[订单_new.采购数量1!=0]
def stock_cl_paixu(stock,Order,Order_column,method='sanhuo'):
    #订单匹配库存数量
    stock = stock.sort_values(['商品编码', '到期日期'], ascending=[True, True])
    stock['排序']=stock.groupby('商品编码')['到期日期'].rank(method='first')
    stock.排序=stock.排序.astype(int).astype(str)
    stock['匹配列'] = stock.商品编码 + stock.排序
    x = stock.排序.astype(int).max()
    for i3 in numpy.arange(1, x + 1):
        Order[i3] = i3
        Order[i3] = Order[i3].astype(str)
        Order[i3] = Order['商品编码'] + Order[i3]
    z = 1
    while z <= x:
        Order = pandas.merge(left=Order,
                      right=stock.loc[:,
                            ['匹配列',
                             '可用数量']],
                      left_on=z,
                      right_on='匹配列',
                      how='left').fillna(0)
        Order = Order.drop('匹配列', axis=1)
        Order.可用数量 = Order.可用数量.astype(int)
        Order = Order.rename(columns={'可用数量': x + z})
        z = z + 1
    订单_bak = Order.reindex(columns=Order_column)
    if method=='zhengxiang':
        Order['采购数量1'] = (Order['采购数量1'] / Order.箱规.fillna(1)
                       ).astype(int) * Order.箱规.fillna(1)
    ix = 1
    订单处理h = pandas.DataFrame()
    while ix <= x and len(Order) > 0:
        订单处理 = Order.sort_values(['商品编码', '采购数量1'], ascending=[True, True])
        订单处理['累加和'] = 订单处理.groupby('商品编码')['采购数量1'].cumsum()
        订单处理[x * 2 + ix] = 订单处理[x + ix] - 订单处理['累加和']
        订单处理.loc[订单处理[x * 2 + ix] >= 0, x * 3 + 1] = 订单处理['采购数量1']
        订单处理.loc[(订单处理[x * 2 + ix] < 0) & (订单处理.采购数量1 + 订单处理[x * 2 + ix]
                                           > 0), x * 3 + 1] = 订单处理.采购数量1 + 订单处理[x * 2 + ix]
        订单处理.loc[(订单处理[x * 2 + ix] < 0) & (订单处理.采购数量1 +
                                           订单处理[x * 2 + ix] <= 0), x * 3 + 1] = 0
        订单处理1 = 订单处理[订单处理[x * 3 + 1] >= 0]
        订单处理1.loc[订单处理1[x * 3 + 1] > 0, '调拨条码'] = 订单处理1[ix]
        订单处理h = 订单处理h.append(订单处理1)
        Order = 订单处理[(订单处理[x * 3 + 1] == 0) | (订单处理[x * 3 + 1] - 订单处理.采购数量1 < 0)]
        Order.采购数量1 = Order.采购数量1 - Order[x * 3 + 1]
        ix = ix + 1
    订单处理_最终 = pandas.merge(
        left=订单处理h,
        right=stock.drop('箱规',axis=1),
        left_on='调拨条码',
        right_on='匹配列',
        how='left')
    订单处理_最终 = 订单处理_最终.reindex(
        columns=[
            '订单号',
            '分配机构',
            '仓库',
            '详细地址',
            '联系人',
            '联系方式',
            '商品编码_x',
            '条码',
            '商品名称_x',
            '采购数量',
            '采购金额',
            '库位',
            '生产批号',
            '商品代码',
            x * 3 + 1,
            '到期日期',
            '箱规',
            '箱数',
            '保质期',
            '体积',
            '重量'])
    订单处理_最终 = 订单处理_最终.rename(
        columns={
            '商品编码_x': '商品编码',
            '商品名称_x': '商品名称',
            x * 3 + 1: '仓库实发数量'})
    订单处理_最终1=订单处理_最终##测试.loc[订单处理_最终.仓库实发数量 != 0]
    return 订单处理_最终1,Order_new(订单处理_最终, 订单_bak)



def data_openpyxl_mingxi(data,path):
    data=data.reset_index()
    wb= openpyxl.load_workbook(path)
    ws=wb['发货明细']
    list_change=['商品名称','采购数量','采购金额','条码']
    for i in ['订单号','分配机构','仓库','详细地址','联系人','联系方式','订单箱数','商品编码']:
        for i2 in excel_cl.merge_cell(data, "订单号", i, list_change):
            ws.merge_cells(i2)
    font = Font(name=u'微软雅黑', size=9)
    font_text = Font(name=u'微软雅黑', size=10,bold=True,color='FFFFFFFF')
    border = Border(left=Side(border_style='thin',color='FF000000'),
    right=Side(border_style='thin',color='FF000000'),
    top=Side(border_style='thin',color='FF000000'),
    bottom=Side(border_style='thin',color='FF000000'))
    fill = PatternFill("solid", fgColor="136E63")
    for j in ws.rows:    # we.rows 获取每一行数据
        for n in j:
            n.alignment =Alignment(horizontal='center', vertical='center',wrapText=True)
            n.font =font
            n.border = border
    for i in excel_cl.reduce_excel_col_name(len(data.columns)):
        ws.column_dimensions[i].width = 9
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['L'].width = 13
    ws.column_dimensions['O'].width = 13
    ws.column_dimensions['P'].width = 13
    ws.column_dimensions['Q'].width = 13
    ws.column_dimensions['N'].width = 13
    for i in range(1,len(data)+2):
        ws.row_dimensions[i].height = 15
    for i in excel_cl.reduce_excel_col_name(len(data.columns)):
        ws[i+str(1)].font =font_text
        ws[i+str(1)].fill = fill
    for i in  range(2,len(data)+2):
        ws['I'+str(i)].alignment = Alignment(horizontal='left', vertical='center')
    wb.save(path)
    print('\033[1;32m 正常:发货明细表已格式化 \033[0m')

def data_openpyxl_huizong(data,path):
    wb= openpyxl.load_workbook(path)
    ws=wb['订单汇总']
    font = Font(name=u'微软雅黑', size=9)
    font_text = Font(name=u'微软雅黑', size=10,bold=True,color='FFFFFF')
    border = Border(left=Side(border_style='thin',color='000000'),
    right=Side(border_style='thin',color='000000'),
    top=Side(border_style='thin',color='000000'),
    bottom=Side(border_style='thin',color='000000'))
    fill = PatternFill("solid", fgColor="136E63")
    for j in ws.rows:    # we.rows 获取每一行数据
        for n in j:
            n.alignment =Alignment(horizontal='left', vertical='center')
            n.font =font
            n.border = border
    for i in excel_cl.reduce_excel_col_name(len(data.columns)):
        ws.column_dimensions[i].width = 12
    for i in range(1,len(data)+2):
        ws.row_dimensions[i].height = 15
    for i in excel_cl.reduce_excel_col_name(len(data.columns)):
        ws[i+str(1)].font =font_text
        ws[i+str(1)].fill = fill
        ws[i+str(1)].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(path)
    print('\033[1;32m 正常:发货汇总表已格式化 \033[0m')

def export_data(data,path,Order_sum,sku):
    data = data.drop_duplicates()
    data.订单号 = data.订单号.astype(str)
    data['数据'] = data.订单号 + data.商品编码
    data = pandas.merge(
        left=data,
        right=data.数据.value_counts().reset_index(),
        left_on='数据',
        right_on='index',
        how='left')
    data = data.loc[(data.仓库实发数量 != 0) | (
        (data.数据_y == 1) & (data.仓库实发数量 == 0))]
    data = pandas.merge(left=data,
                       right=sku.loc[:,
                             ['公司货号',
                              '茂浦采购价未税',
                              ]].drop_duplicates(),
                       left_on='商品代码',
                       right_on='公司货号',
                       how='left')
    data.loc[data.箱规 == 0, '箱规'] = 1
    data.loc[data.仓库实发数量 != 0, '箱数'] = data.loc[data.仓库实发数量 !=
                                                         0, '仓库实发数量'] / (data.loc[data.仓库实发数量 != 0, '箱规'].fillna(1))
    data.loc[data.仓库实发数量 != 0,
                '箱数'] = data.loc[data.仓库实发数量 != 0,
                                    '箱数'].map(lambda x: math.ceil(x))
    data['总体积'] = data.体积 * data.箱数
    data['总重量'] = data.重量 * data.箱数
    data['货品总额'] = data.采购金额 / 0.63 / data.采购数量 * 0.56 * data.仓库实发数量
    订单处理_订单汇总 = data.pivot_table(
        index=[
            '订单号',
            '分配机构',
            '仓库',
            '详细地址',
            '联系方式'],
        values=[
            '仓库实发数量',
            '箱数',
            '总体积',
            '货品总额',
            '总重量'],
        aggfunc='sum').reset_index()
    订单处理_订单汇总['填开日期'] = datetime.today().strftime('%Y-%m-%d')
    订单处理_订单汇总['客户名称'] = '上海茂浦电子商务有限公司'
    订单处理_订单汇总['始发地'] = '上海'
    订单处理_订单汇总['电商名称'] = '京东商城'
    订单处理_订单汇总 = 订单处理_订单汇总.reindex(
        columns=[
            '填开日期',
            '客户名称',
            '始发地',
            '分配机构',
            '电商名称',
            '订单号',
            '仓库实发数量',
            '箱数',
            '货品总额',
            '总重量',
            '总体积',
            '预约状态',
            '预约日期',
            '送货时间段',
            '预约号',
            '仓库',
            '详细地址',
            '联系方式',
            '备注'])
    订单处理_订单汇总 = 订单处理_订单汇总.sort_values(['分配机构', '仓库', '订单号'])
    订单处理_发货明细 = data.sort_values(['分配机构', '仓库', '订单号'])
    订单处理_发货明细.loc[订单处理_发货明细.到期日期.notnull(),
                  '有效期'] = 订单处理_发货明细.loc[订单处理_发货明细.到期日期.notnull(),
                                         '到期日期'].map(lambda x: x.strftime('%Y-%m-%d'))
    订单处理_发货明细 = 订单处理_发货明细.drop('到期日期', axis=1)
    # 标记
    订单处理_发货明细1 = 订单处理_发货明细.reindex(
        columns=[
            '订单号',
            '分配机构',
            '仓库',
            '详细地址',
            '联系人',
            '联系方式',
            '商品编码',
            '条码',
            '茂浦采购价未税',
            '商品名称',
            '采购数量',
            '采购金额',
            '库位',
            '生产批号',
            '商品代码',
            '仓库实发数量',
            '有效期',
            '箱数',
            '箱规',
            '保质期',
            '体积',
            '重量',
            '总体积',
            '总重量'])
    订单处理_发货明细 = 订单处理_发货明细.reindex(
        columns=[
            '订单号',
            '分配机构',
            '仓库',
            '详细地址',
            '联系人',
            '联系方式',
            '商品编码',
            '条码',
            '商品名称',
            '采购数量',
            '采购金额',
            '库位',
            '生产批号',
            '商品代码',
            '仓库实发数量',
            '有效期',
            '箱数',
            '箱规',
            '保质期',
            '总体积',
            '总重量'])
    订单处理_发货明细.条码 = 订单处理_发货明细.条码.astype(str)
    订单处理_发货明细 = 订单处理_发货明细.sort_values(['分配机构', '仓库', '订单号', '商品编码'])
    订单箱数 = 订单处理_发货明细.pivot_table(
        index='订单号',
        values='箱数',
        aggfunc='sum').reset_index()
    订单箱数 = 订单箱数.rename(columns={'箱数': '订单箱数'})
    订单处理_发货明细 = pandas.merge(left=订单处理_发货明细, right=订单箱数, on='订单号', how='left')
    订单处理_发货明细1 = pandas.merge(left=订单处理_发货明细1, right=订单箱数, on='订单号', how='left')
    # 添加订单箱数
    订单处理_发货明细 = 订单处理_发货明细.set_index(['订单号',
                                     '分配机构',
                                     '仓库',
                                     '详细地址',
                                     '联系人',
                                     '联系方式',
                                     '订单箱数',
                                     '商品编码',
                                     '商品名称',
                                     '采购数量',
                                     '采购金额',
                                     '条码',
                                     '库位',
                                     '生产批号',
                                     '商品代码'])
    定义 = data.保质期.fillna(0).map(lambda x: timedelta(x))
    data['生产日期'] = (data.到期日期 - 定义 + 定义 / 3).map(lambda x: x.date())
    data['是否超保'] = '否'
    data.loc[data.生产日期 < datetime.now().date() + timedelta(5), '是否超保'] = '超保'
    导出路径 = pandas.ExcelWriter(path + '出库单(发仓库).xlsx')
    订单处理_订单汇总.to_excel(导出路径, '订单汇总', index=None)
    订单处理_发货明细.to_excel(导出路径, '发货明细')
    if 订单处理_发货明细.reset_index().loc[:,['订单号','商品编码','采购数量']].drop_duplicates().采购数量.sum()-Order_sum==0:
        print('\033[1;32m 正常:订单数量和出库订单数量一致 \033[0m')
    else:
        print('\033[1;31m 警告:订单数量和出库订单数量不一致 \033[0m')
    data['生产日期'] = (data.到期日期 - 定义).map(lambda x: x.date())
    订单处理_超保 = data[(data.是否超保 == '超保') & (
        data.仓库实发数量 != 0) & (data.仓库实发数量.notnull())]
    data.loc[data.是否超保 == '超保']
    订单处理_超保['采销部门'] = '个护'
    订单处理_超保明细 = 订单处理_超保.reindex(columns=['分配机构',
                                         '仓库',
                                         '采销部门',
                                         '订单号',
                                         '商品编码',
                                         '商品名称',
                                         '仓库实发数量',
                                         '保质期',
                                         '生产日期',
                                         '到期日期']).sort_values(['订单号',
                                                               '分配机构',
                                                               '仓库'])
    订单处理_超保明细['到期日期'] = 订单处理_超保明细.到期日期.map(lambda x: x.strftime('%Y-%m-%d'))
    订单处理_超保明细.to_excel(导出路径, '超保明细', index=None)
    data_out = 订单处理_发货明细1.reset_index()
    data_out['标卡'] = data_out.订单号 + data_out.商品编码

    data_out1 = data_out.fillna('dummy').groupby(['订单号', '商品编码', '商品名称', '订单箱数', '茂浦采购价未税', '箱规', '体积', '重量'])[
        '仓库实发数量', '箱数', '总体积', '总重量'].sum().reset_index().replace('dummy',numpy.nan)
    data_out.仓库实发数量 = data_out.仓库实发数量.astype(numpy.int64).astype(str)
    data_out.箱数 = data_out.箱数.fillna(0).astype(numpy.int64).astype(str)
    data_out.生产批号 = data_out.生产批号.fillna(0).astype(str)
    data_all = pandas.DataFrame()
    for i in data_out.标卡.unique():
        data1 = data_out.loc[data_out.标卡 == i]
        data1['生产批号cat'] = data1.生产批号.str.cat(sep=';')
        data1['条码cat'] = data1.条码.str.cat(sep=';')
        data1['库位cat'] = data1.库位.astype(str).str.cat(sep=';')
        data1['商品代码cat'] = data1.商品代码.str.cat(sep=';')
        data1['仓库实发数量cat'] = data1.仓库实发数量.str.cat(sep=';')
        data1['有效期cat'] = data1.有效期.str.cat(sep=';')
        data1['箱数cat'] = data1.箱数.str.cat(sep=';')
        data1 = data1.drop(['订单箱数',
                            '商品名称',
                            '条码',
                            '生产批号',
                            '库位',
                            '商品代码',
                            '仓库实发数量',
                            '有效期',
                            '箱数',
                            '体积',
                            '重量',
                            '标卡',
                            '总体积',
                            '总重量',
                            '茂浦采购价未税',
                            '箱规'],
                           axis=1)
        data_all = data_all.append(data1)
    data_out2 = data_all.drop('index', axis=1).drop_duplicates()
    data_out3 = pandas.merge(data_out2, data_out1, on=['订单号', '商品编码'], how='left')
    data_out3['始发仓库'] = '华夏龙'
    data_out3['下单日期'] = datetime.today()
    data_out3.下单日期 = data_out3.下单日期.map(lambda x: x.strftime('%Y-%m-%d'))
    #data_out3.茂浦采购价未税 = data_out3.茂浦采购价未税 * 1.13
    data_out3['出库折扣'] = 0.63
    data_out3 = data_out3.reindex(
        columns=[
            '订单号',
            '商品代码cat',
            '商品编码',
            '条码cat',
            '茂浦采购价未税',
            '商品名称',
            '箱规',
            '始发仓库',
            '分配机构',
            '下单日期',
            '',
            '库位cat',
            '生产批号cat',
            '采购数量',
            '',
            '',
            '',
            '',
            '仓库实发数量cat',
            '仓库实发数量',
            '',
            '',
            '',
            '有效期cat',
            '箱数cat',
            '箱数',
            '订单箱数',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '出库折扣',
            '',
            '',
            '',
            '体积',
            '重量',
            '总体积',
            '总重量',
            '保质期'])
    data_out3.to_excel(导出路径, '出库明细', index=None)
    导出路径.save()
    data_openpyxl_mingxi(订单处理_发货明细, path + '出库单(发仓库).xlsx')
    data_openpyxl_huizong(订单处理_订单汇总,path + '出库单(发仓库).xlsx')
    print('\033[1;32m 正常:导出完成 \033[0m')
def main():
    pass
if __name__ == '__main__':
    main()
